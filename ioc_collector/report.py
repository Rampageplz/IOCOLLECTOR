"""Geração e exibição de relatórios de IOCs."""

from __future__ import annotations

import argparse
import csv
import datetime
import json
import logging
from collections import Counter, defaultdict
from dataclasses import asdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import requests
import xlwt
import pandas as pd

from rich.console import Console
from rich.table import Table

from .models import Report

ALERTS_FILE = Path(__file__).resolve().parent / "alerts.json"
EXPECTED_FEEDS = ["AbuseIPDB", "OTX", "URLHaus"]


def _load_alerts() -> List[Dict[str, any]]:
    if not ALERTS_FILE.exists():
        raise FileNotFoundError("alerts.json não encontrado")
    with ALERTS_FILE.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def _filter_alerts(
    alerts: Iterable[Dict[str, any]],
    date: Optional[str] = None,
    ioc_type: Optional[str] = None,
    source: Optional[str] = None,
    value: Optional[str] = None,
) -> List[Dict[str, any]]:
    result = []
    for a in alerts:
        if date and a.get("date") != date:
            continue
        if ioc_type and a.get("ioc_type") != ioc_type:
            continue
        if source and a.get("source") != source:
            continue
        if value and a.get("ioc_value") != value:
            continue
    result.append(a)
    return result


def build_correlation_dataframe(iocs: List[Dict[str, any]]):
    """Return DataFrame with correlation columns based on IOC value."""
    df = pd.DataFrame(iocs)
    if df.empty:
        return df

    grouped = df.groupby("ioc_value")
    sources = grouped["source"].apply(lambda x: sorted(set(x))).reset_index(name="sources")
    unique = grouped.first().reset_index()
    merged = unique.merge(sources, on="ioc_value")
    merged["source_count"] = merged["sources"].apply(len)
    merged["risk_score"] = merged["source_count"] * 10
    return merged.drop(columns=["source"], errors="ignore")


def _fetch_ip_info(ip: str) -> tuple[str, str]:
    """Return (country, ASN) information for an IP using ip-api.com."""
    try:
        resp = requests.get(
            f"http://ip-api.com/json/{ip}?fields=status,country,as",
            timeout=10,
        )
        data = resp.json()
        if data.get("status") == "success":
            return data.get("country", ""), data.get("as", "")
    except Exception:
        logging.exception("Erro ao consultar ip-api para %s", ip)
    return "", ""


def _mitigation_ip(count: int) -> str:
    if count >= 3:
        return "Bloquear IP imediatamente"
    if count == 2:
        return "Monitorar atividade do IP"
    return "Analisar manualmente"


def _mitigation_url(count: int) -> str:
    return "Bloquear via proxy" if count >= 2 else "Monitorar acessos"


def _mitigation_hash(count: int) -> str:
    return "Quarentenar arquivo" if count >= 2 else "Investigar"


def _mitigation_domain(count: int) -> str:
    return "Bloquear dominio" if count >= 2 else "Monitorar"


def _prepare_ip_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(
            columns=[
                "IP",
                "Fontes",
                "País",
                "ASN",
                "Data",
                "Quantidade de Fontes",
                "Score de Risco",
                "Mitigação recomendada",
            ]
        )
    countries = []
    asns = []
    for ip in df["ioc_value"]:
        country, asn = _fetch_ip_info(ip)
        countries.append(country)
        asns.append(asn)
    df = df.copy()
    df["Fontes"] = df["sources"].apply(lambda s: ", ".join(s))
    df["Data"] = df["date"]
    df["País"] = countries
    df["ASN"] = asns
    df["Quantidade de Fontes"] = df["source_count"]
    df["Score de Risco"] = df["risk_score"]
    df["Mitigação recomendada"] = df["source_count"].apply(_mitigation_ip)
    df.rename(columns={"ioc_value": "IP"}, inplace=True)
    return df[
        [
            "IP",
            "Fontes",
            "País",
            "ASN",
            "Data",
            "Quantidade de Fontes",
            "Score de Risco",
            "Mitigação recomendada",
        ]
    ]


def _prepare_url_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(
            columns=[
                "URL",
                "Fontes",
                "Descrição da ameaça",
                "Data",
                "Quantidade de Fontes",
                "Score de Risco",
                "Mitigação recomendada",
            ]
        )
    df = df.copy()
    df["Fontes"] = df["sources"].apply(lambda s: ", ".join(s))
    df["Data"] = df["date"]
    df["Quantidade de Fontes"] = df["source_count"]
    df["Score de Risco"] = df["risk_score"]
    df["Mitigação recomendada"] = df["source_count"].apply(_mitigation_url)
    df.rename(columns={"ioc_value": "URL", "description": "Descrição da ameaça"}, inplace=True)
    return df[
        [
            "URL",
            "Fontes",
            "Descrição da ameaça",
            "Data",
            "Quantidade de Fontes",
            "Score de Risco",
            "Mitigação recomendada",
        ]
    ]


def _prepare_hash_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(
            columns=[
                "Hash",
                "Tipo do hash",
                "Fontes",
                "Data",
                "Quantidade de Fontes",
                "Score de Risco",
                "Mitigação recomendada",
            ]
        )
    df = df.copy()
    df["Fontes"] = df["sources"].apply(lambda s: ", ".join(s))
    df["Data"] = df["date"]
    df["Quantidade de Fontes"] = df["source_count"]
    df["Score de Risco"] = df["risk_score"]
    df["Mitigação recomendada"] = df["source_count"].apply(_mitigation_hash)
    df.rename(columns={"ioc_value": "Hash", "ioc_type": "Tipo do hash"}, inplace=True)
    return df[
        [
            "Hash",
            "Tipo do hash",
            "Fontes",
            "Data",
            "Quantidade de Fontes",
            "Score de Risco",
            "Mitigação recomendada",
        ]
    ]


def _prepare_domain_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(
            columns=[
                "Domínio",
                "Fontes",
                "Categoria da ameaça",
                "Data",
                "Quantidade de Fontes",
                "Score de Risco",
                "Mitigação recomendada",
            ]
        )
    df = df.copy()
    df["Fontes"] = df["sources"].apply(lambda s: ", ".join(s))
    df["Data"] = df["date"]
    df["Quantidade de Fontes"] = df["source_count"]
    df["Score de Risco"] = df["risk_score"]
    df["Mitigação recomendada"] = df["source_count"].apply(_mitigation_domain)
    df.rename(columns={"ioc_value": "Domínio", "description": "Categoria da ameaça"}, inplace=True)
    return df[
        [
            "Domínio",
            "Fontes",
            "Categoria da ameaça",
            "Data",
            "Quantidade de Fontes",
            "Score de Risco",
            "Mitigação recomendada",
        ]
    ]


def save_correlation_reports(iocs: List[Dict[str, any]], csv_path: Path, xlsx_path: Path) -> None:
    """Save correlation report to CSV and Excel with separated IOC sheets."""
    df = build_correlation_dataframe(iocs)
    if df.empty:
        logging.info("Nenhum dado para o relatorio de correlacao")
        return

    df.to_csv(csv_path, index=False)

    writer = pd.ExcelWriter(xlsx_path, engine="openpyxl")

    ip_mask = df["ioc_type"].str.contains("ip", case=False, na=False)
    url_mask = df["ioc_type"].str.contains("url", case=False, na=False)
    hash_mask = df["ioc_type"].str.contains("hash", case=False, na=False) | df["ioc_type"].str.contains("sha", case=False, na=False) | df["ioc_type"].str.contains("md5", case=False, na=False)
    domain_mask = df["ioc_type"].str.contains("domain", case=False, na=False)

    logging.info("Gerando aba de IPs")
    _prepare_ip_df(df[ip_mask]).to_excel(writer, sheet_name="IPs", index=False)
    logging.info("Gerando aba de URLs")
    _prepare_url_df(df[url_mask]).to_excel(writer, sheet_name="URLs", index=False)
    logging.info("Gerando aba de Hashes")
    _prepare_hash_df(df[hash_mask]).to_excel(writer, sheet_name="Hashes", index=False)
    logging.info("Gerando aba de Domínios")
    _prepare_domain_df(df[domain_mask]).to_excel(writer, sheet_name="Domínios", index=False)

    writer.close()
    logging.info(
        "Relatorio de correlacao salvo em %s e %s",
        csv_path.resolve(),
        xlsx_path.resolve(),
    )


def generate_report(
    date: str,
    *,
    ioc_type: Optional[str] = None,
    source: Optional[str] = None,
    value: Optional[str] = None,
    top_count: int = 10,
    all_history: bool = False,
    sort: bool = False,
) -> Report:
    alerts = _load_alerts()
    if all_history:
        daily = _filter_alerts(alerts, ioc_type=ioc_type, source=source, value=value)
    else:
        daily = _filter_alerts(alerts, date=date, ioc_type=ioc_type, source=source, value=value)

    if sort:
        daily.sort(key=lambda x: x.get("lastReportedAt") or x.get("date"))

    if not daily:
        raise ValueError(f"Nenhum IOC encontrado para a data {date}")

    by_source = Counter(a.get("source") for a in daily)
    by_type = Counter(a.get("ioc_type") for a in daily)
    total = len(daily)

    occurrences: Dict[str, List[str]] = defaultdict(list)
    for a in daily:
        occurrences[a.get("ioc_value")].append(a.get("source"))

    duplicates = {
        val: sorted(set(srcs)) for val, srcs in occurrences.items() if len(set(srcs)) > 1
    }

    top_values = Counter([a.get("ioc_value") for a in daily]).most_common(top_count)
    coverage = {src: round((cnt / total) * 100, 2) for src, cnt in by_source.items()}
    missing_feeds = [f for f in EXPECTED_FEEDS if f not in by_source]
    insights = []
    if by_source:
        main_feed, main_cnt = max(by_source.items(), key=lambda x: x[1])
        perc = coverage[main_feed]
        insights.append(f"{main_feed} representa {perc:.0f}% dos IOCs de {date}.")
    if duplicates:
        insights.append(f"{len(duplicates)} IOCs aparecem em mais de um feed.")
    else:
        insights.append("Nenhum IOC duplicado encontrado.")

    return Report(
        date=date,
        total_iocs=len(daily),
        by_source=dict(by_source),
        by_type=dict(by_type),
        duplicates=duplicates,
        top_values=top_values,
        iocs=daily,
        coverage=coverage,
        missing_feeds=missing_feeds,
        insights=insights,
    )


def _save_json(report: Report, path: Path) -> None:
    path.write_text(json.dumps(asdict(report), indent=2, ensure_ascii=False))
    logging.info("✅ Relatório JSON salvo em %s", path.resolve())


def _save_csv(report: Report, path: Path) -> None:
    if not report.iocs:
        raise ValueError("Relatório vazio")
    fieldnames = sorted({k for item in report.iocs for k in item.keys()})
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for item in report.iocs:
            writer.writerow({fn: item.get(fn, "") for fn in fieldnames})
    logging.info("✅ CSV salvo em %s", path.resolve())


def _save_txt(report: Report, path: Path) -> None:
    table = _build_table(report)
    console = Console(record=True)
    console.print(table)
    path.write_text(console.export_text())
    logging.info("✅ TXT salvo em %s", path.resolve())


def _save_pdf(report: Report, path: Path) -> None:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, f"Relatório {report.date}", ln=True, align="C")

    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, "Resumo por Feed", ln=True)
    for src, total in report.by_source.items():
        pdf.cell(0, 8, f"{src}: {total}", ln=True)
    pdf.cell(0, 8, f"Total: {report.total_iocs}", ln=True)

    pdf.ln(4)
    pdf.cell(0, 10, "Top IOCS", ln=True)
    for val, cnt in report.top_values:
        pdf.cell(0, 8, f"{val} - {cnt}", ln=True)

    if report.duplicates:
        pdf.ln(4)
        pdf.cell(0, 10, "Duplicados", ln=True)
        for val, feeds in report.duplicates.items():
            pdf.cell(0, 8, f"{val}: {', '.join(feeds)}", ln=True)

    pdf.output(str(path))
    logging.info("✅ PDF salvo em %s", path.resolve())


def _save_xls(report: Report, path: Path) -> None:
    wb = xlwt.Workbook()
    bold = xlwt.easyxf("font: bold on; pattern: pattern solid, fore_colour gray25")

    ws_summary = wb.add_sheet("Resumo")
    ws_summary.write(0, 0, "Feed", bold)
    ws_summary.write(0, 1, "Quantidade", bold)
    row = 1
    for src, total in report.by_source.items():
        ws_summary.write(row, 0, src)
        ws_summary.write(row, 1, total)
        row += 1
    ws_summary.write(row, 0, "Total", bold)
    ws_summary.write(row, 1, report.total_iocs)

    ws_type = wb.add_sheet("Por Tipo")
    ws_type.write(0, 0, "Tipo", bold)
    ws_type.write(0, 1, "Quantidade", bold)
    row = 1
    for t, count in report.by_type.items():
        ws_type.write(row, 0, t)
        ws_type.write(row, 1, count)
        row += 1

    ws_dups = wb.add_sheet("Duplicados")
    ws_dups.write(0, 0, "IOC", bold)
    ws_dups.write(0, 1, "Feeds", bold)
    row = 1
    for val, feeds in report.duplicates.items():
        ws_dups.write(row, 0, val)
        ws_dups.write(row, 1, ", ".join(feeds))
        row += 1

    ws_top = wb.add_sheet("Top")
    ws_top.write(0, 0, "IOC", bold)
    ws_top.write(0, 1, "Total", bold)
    row = 1
    for val, cnt in report.top_values:
        ws_top.write(row, 0, val)
        ws_top.write(row, 1, cnt)
        row += 1

    ws_iocs = wb.add_sheet("IOC List")
    header = sorted({k for item in report.iocs for k in item.keys()}) if report.iocs else []
    for col, h in enumerate(header):
        ws_iocs.write(0, col, h, bold)
        ws_iocs.col(col).width = 256 * 20
    for row_idx, item in enumerate(report.iocs, start=1):
        for col_idx, h in enumerate(header):
            val = item.get(h, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            ws_iocs.write(row_idx, col_idx, val)

    wb.save(str(path))
    logging.info("✅ XLS salvo em %s", path.resolve())


def _save_xlsx(report: Report, path: Path) -> None:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Resumo"
    ws_summary.append(["Feed", "Quantidade"])
    for src, total in report.by_source.items():
        ws_summary.append([src, total])
    ws_summary.append(["Total", report.total_iocs])

    ws_type = wb.create_sheet("Por Tipo")
    ws_type.append(["Tipo", "Quantidade"])
    for t, count in report.by_type.items():
        ws_type.append([t, count])

    ws_dups = wb.create_sheet("Duplicados")
    ws_dups.append(["IOC", "Feeds"])
    for val, feeds in report.duplicates.items():
        ws_dups.append([val, ", ".join(feeds)])

    ws_top = wb.create_sheet("Top")
    ws_top.append(["IOC", "Total"])
    for val, cnt in report.top_values:
        ws_top.append([val, cnt])

    ws_iocs = wb.create_sheet("IOC List")
    if not report.iocs:
        header = []
    else:
        header = sorted({k for item in report.iocs for k in item.keys()})
    ws_iocs.append(header)
    ws_iocs.freeze_panes = "A2"
    last_col = get_column_letter(len(header)) if header else "A"
    ws_iocs.auto_filter.ref = f"A1:{last_col}1"
    for item in report.iocs:
        row = []
        for h in header:
            val = item.get(h, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            row.append(val)
        ws_iocs.append(row)

    for col in ws_iocs.columns:
        ws_iocs.column_dimensions[col[0].column_letter].width = 20

    wb.save(path)
    logging.info("✅ XLSX salvo em %s", path.resolve())


def _build_table(report: Report) -> Table:
    table = Table(title=f"Resumo {report.date}")
    table.add_column("Feed")
    table.add_column("Quantidade", justify="right")
    table.add_column("%", justify="right")
    for src, total in report.by_source.items():
        perc = f"{report.coverage.get(src, 0):.1f}%"
        table.add_row(src, str(total), perc)
    table.add_row("Total", str(report.total_iocs), "100%")
    if report.missing_feeds:
        table.caption = "Ausentes: " + ", ".join(report.missing_feeds)
    return table


def print_report(report: Report, *, only_duplicates: bool = False, only_top: bool = False) -> None:
    console = Console()
    if not only_duplicates and not only_top:
        table = _build_table(report)
        console.print(table)

    if not only_top:
        if report.duplicates:
            dup_table = Table(title="Duplicados")
            dup_table.add_column("IOC")
            dup_table.add_column("Feeds")
            for val, feeds in report.duplicates.items():
                dup_table.add_row(val, ", ".join(feeds))
            console.print(dup_table)
        else:
            console.print("[cyan]Sem correlacao entre os feeds.[/cyan]")

    if not only_duplicates:
        counts = [c for _, c in report.top_values]
        if counts and any(c > 1 for c in counts):
            top_table = Table(title="Top")
            top_table.add_column("IOC")
            top_table.add_column("Total", justify="right")
            for val, count in report.top_values:
                if count > 1:
                    top_table.add_row(val, str(count))
            console.print(top_table)
        elif only_top:
            console.print("[cyan]Nenhum IOC recorrente encontrado.[/cyan]")


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera relatório de IOCs")
    parser.add_argument(
        "--date",
        default=datetime.date.today().isoformat(),
        help="Data dos IOCs (YYYY-MM-DD)",
    )
    parser.add_argument("--output-json", help="Salvar relatório em JSON")
    parser.add_argument("--output-csv", help="Salvar relatório em CSV")
    parser.add_argument("--output-txt", help="Salvar relatório em TXT")
    parser.add_argument("--output-xls", help="Salvar relatório em XLS")
    parser.add_argument("--output-xlsx", help="Salvar relatório em Excel")
    parser.add_argument("--output-pdf", help="Salvar relatório em PDF")
    parser.add_argument("--type", dest="ioc_type", help="Filtrar por tipo de IOC")
    parser.add_argument("--source", help="Filtrar por feed específico")
    parser.add_argument("--value", help="Filtrar por IOC específico")
    parser.add_argument("--top-count", type=int, default=10)
    parser.add_argument("--all", action="store_true", help="Usar todo o histórico")
    parser.add_argument("--sort", action="store_true", help="Ordenar por data/hora")
    parser.add_argument("--only-duplicates", action="store_true", help="Mostrar apenas duplicados")
    parser.add_argument("--only-top", action="store_true", help="Mostrar apenas a seção Top")
    args = parser.parse_args()

    try:
        report = generate_report(
            args.date,
            ioc_type=args.ioc_type,
            source=args.source,
            value=args.value,
            top_count=args.top_count,
            all_history=args.all,
            sort=args.sort,
        )
    except ValueError as exc:
        Console().print(f"[yellow]⚠️ {exc}[/yellow]")
        return

    print_report(report, only_duplicates=args.only_duplicates, only_top=args.only_top)

    if args.output_json:
        _save_json(report, Path(args.output_json))
    if args.output_csv:
        _save_csv(report, Path(args.output_csv))
    if args.output_txt:
        _save_txt(report, Path(args.output_txt))
    if args.output_xls:
        _save_xls(report, Path(args.output_xls))
    if args.output_xlsx:
        _save_xlsx(report, Path(args.output_xlsx))
    if args.output_pdf:
        _save_pdf(report, Path(args.output_pdf))


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    main()
